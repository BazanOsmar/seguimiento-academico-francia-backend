import re

import openpyxl
from django.db import transaction

from .models import Estudiante

_PATRON_HOJA       = re.compile(r'^\S+\s+[A-Za-z]$')
_COLUMNAS_ESPERADAS = ['PATERNO', 'MATERNO', 'NOMBRES']


def crear_estudiante_solo(datos):
    """
    Crea un Estudiante sin tutor, generando el identificador automáticamente
    con el patrón: {ini_paterno}{ini_materno}{ini_nombre}{grado}{paralelo}-{nro:02d}
    """
    curso      = datos['curso']
    ap_paterno = datos.get('apellido_paterno', '').strip().upper()
    ap_materno = datos.get('apellido_materno', '').strip().upper()
    nombre     = datos['nombre'].strip().upper()

    base_nro = Estudiante.objects.filter(curso=curso).count() + 1
    nro = base_nro
    while True:
        cand = _generar_identificador(ap_paterno, ap_materno, nombre, curso.grado, curso.paralelo, nro)
        if not Estudiante.objects.filter(identificador=cand).exists():
            break
        nro += 1

    return Estudiante.objects.create(
        nombre=nombre,
        apellido_paterno=ap_paterno,
        apellido_materno=ap_materno,
        identificador=cand,
        curso=curso,
        activo=True,
    )



# ── Importación desde Excel ───────────────────────────────────────

def _parsear_hoja(nombre_hoja: str) -> tuple:
    partes = nombre_hoja.strip().split()
    return partes[0], partes[1]


def _generar_identificador(ap_paterno, ap_materno, nombre, grado, paralelo, nro_lista):
    ini_ap  = ap_paterno[0].upper()        if ap_paterno else "X"
    ini_am  = ap_materno[0].upper()        if ap_materno else "X"
    ini_nom = nombre.split()[0][0].upper() if nombre     else "X"
    grado_paralelo = (grado + paralelo).upper().replace(" ", "")
    return f"{ini_ap}{ini_am}{ini_nom}{grado_paralelo}-{str(nro_lista).zfill(2)}"


def validar_formato_excel(archivo) -> list:
    """
    Valida que el archivo Excel tenga el formato esperado antes de importar.
    Retorna lista de errores (vacía = formato correcto).
    Deja el puntero del archivo al inicio para que pueda usarse después.
    """
    errores = []

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
    except Exception:
        return ["El archivo no es un Excel válido o está dañado."]
    finally:
        archivo.seek(0)

    if not wb.sheetnames:
        return ["El archivo no contiene ninguna hoja."]

    hojas_validas = 0
    for nombre_hoja in wb.sheetnames:
        if not _PATRON_HOJA.match(nombre_hoja.strip()):
            errores.append(
                f"Hoja '{nombre_hoja}': nombre inválido — se esperaba formato como '3ro A'."
            )
            continue

        ws   = wb[nombre_hoja]
        fila3 = [str(c.value).strip().upper() if c.value else '' for c in ws[3]]
        cols  = fila3[1:4]  # columnas 2, 3, 4 (índice 1, 2, 3)

        if cols != _COLUMNAS_ESPERADAS:
            errores.append(
                f"Hoja '{nombre_hoja}': encabezados incorrectos en fila 3. "
                f"Se esperaba PATERNO · MATERNO · NOMBRES — "
                f"se encontró: {' · '.join(c or '(vacío)' for c in cols)}."
            )
            continue

        hojas_validas += 1

    if hojas_validas == 0:
        errores.append("No se encontró ninguna hoja con formato válido.")

    return errores


@transaction.atomic
def importar_estudiantes_desde_excel(archivo) -> dict:
    """
    Procesa un Excel con una hoja por curso (ej: '3ro A').
    Filas 1-3: cabecera. Desde fila 4: nro | ap_paterno | ap_materno | nombre(s).
    Crea los cursos automáticamente si no existen.
    """
    from backend.apps.academics.models import Curso

    wb = openpyxl.load_workbook(archivo, data_only=True)

    importados = 0
    omitidos   = 0
    errores    = []

    existentes = set(Estudiante.objects.values_list('identificador', flat=True))

    for nombre_hoja in wb.sheetnames:
        try:
            grado, paralelo = _parsear_hoja(nombre_hoja)
        except (IndexError, ValueError):
            errores.append(f"Hoja '{nombre_hoja}': nombre inválido.")
            continue

        curso, _ = Curso.objects.get_or_create(grado=grado, paralelo=paralelo)

        for fila in wb[nombre_hoja].iter_rows(min_row=4, values_only=True):
            nro, paterno, materno, nombres = fila[0], fila[1], fila[2], fila[3]

            if not any([nro, paterno, materno, nombres]):
                continue

            nro_lista     = int(nro) if nro else 0
            ap_paterno    = str(paterno).strip().upper() if paterno else ""
            ap_materno    = str(materno).strip().upper() if materno else ""
            nombre_limpio = str(nombres).strip().upper() if nombres else ""

            if not nombre_limpio:
                errores.append(
                    f"Hoja '{nombre_hoja}' fila {nro_lista}: "
                    f"datos incompletos (falta nombre)."
                )
                continue

            identificador = _generar_identificador(
                ap_paterno, ap_materno, nombre_limpio, grado, paralelo, nro_lista,
            )

            if identificador in existentes:
                omitidos += 1
                continue

            Estudiante.objects.create(
                nombre=nombre_limpio,
                apellido_paterno=ap_paterno,
                apellido_materno=ap_materno,
                identificador=identificador,
                curso=curso,
                activo=True,
            )
            existentes.add(identificador)
            importados += 1

    return {"importados": importados, "omitidos": omitidos, "errores": errores}
