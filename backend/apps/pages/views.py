from datetime import datetime

from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404


def login_view(request):
    return render(request, 'auth/login.html')


def page_not_found_view(request, exception=None):
    return render(request, '404.html', status=404)


def director_view(request):
    return render(request, 'director/dashboard.html')


def director_estudiantes_view(request):
    return render(request, 'director/estudiantes.html')


def director_curso_estudiantes_view(request, curso_id):
    from backend.apps.academics.models import Curso
    curso = get_object_or_404(Curso, pk=curso_id)
    return render(request, 'director/curso_estudiantes.html', {
        'curso_id': curso_id,
        'curso_nombre': f"{curso.grado} {curso.paralelo}",
    })


def director_perfil_estudiante_view(request, curso_id, estudiante_id):
    return render(request, 'director/perfil_estudiante.html', {
        'curso_id': curso_id,
        'estudiante_id': estudiante_id,
    })


def director_usuarios_view(request):
    return render(request, 'director/usuarios.html')


def director_perfil_usuario_view(request, user_id):
    return render(request, 'director/perfil_usuario.html', {'user_id': user_id})


def director_asistencia_view(request):
    return render(request, 'director/asistencia.html')


def _autenticar_por_token(request):
    """Valida JWT desde query param ?token= o header Authorization: Bearer."""
    from rest_framework_simplejwt.tokens import AccessToken
    from backend.apps.users.models import User

    # 1. Intentar desde query param
    raw = request.GET.get('token', '').strip()
    # 2. Fallback: header Authorization (para fetchAPI del frontend web)
    if not raw:
        auth = request.META.get('HTTP_AUTHORIZATION', '')
        if auth.startswith('Bearer '):
            raw = auth[7:].strip()
    if not raw:
        return None
    try:
        validated = AccessToken(raw)
        return User.objects.select_related('tipo_usuario').get(pk=validated['user_id'])
    except Exception:
        return None


_ESTADO_LETRA = {'PRESENTE': 'P', 'FALTA': 'F', 'ATRASO': 'A', 'LICENCIA': 'L'}
_DIAS_ES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
_MESES_ES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]


def director_asistencia_exportar_view(request):
    import calendar
    from datetime import date
    from backend.apps.academics.models import Curso
    from backend.apps.attendance.models import AsistenciaSesion, Asistencia
    from backend.apps.students.models import Estudiante

    # ── Autenticación: token JWT en query param ──────────────────
    user = _autenticar_por_token(request)
    if not user or not user.tipo_usuario or user.tipo_usuario.nombre not in ('Director', 'Regente'):
        return JsonResponse({'errores': 'No autorizado.'}, status=403)

    curso_id = request.GET.get('curso_id', '').strip()
    mes_param = request.GET.get('mes', '').strip()  # formato YYYY-MM
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    # Soportar tanto ?mes=YYYY-MM como ?fecha_desde=...&fecha_hasta=...
    if mes_param and not fecha_desde:
        fecha_desde = f"{mes_param}-01"
        try:
            parts = mes_param.split('-')
            y, m = int(parts[0]), int(parts[1])
            _, ld = calendar.monthrange(y, m)
            fecha_hasta = f"{mes_param}-{str(ld).zfill(2)}"
        except (ValueError, IndexError):
            fecha_hasta = fecha_desde

    if not fecha_hasta:
        fecha_hasta = fecha_desde

    ctx = {
        'error': None,
        'curso': None,
        'mes_display': '',
        'dias_habiles': [],
        'estudiantes': [],
        'generado_en': datetime.now().strftime('%d/%m/%Y %H:%M'),
    }

    if not curso_id or not fecha_desde:
        ctx['error'] = 'Faltan parámetros: se requiere curso y mes.'
        return render(request, 'director/asistencia_exportar.html', ctx)

    try:
        curso = Curso.objects.get(pk=int(curso_id))
    except (Curso.DoesNotExist, ValueError):
        ctx['error'] = 'Curso no encontrado.'
        return render(request, 'director/asistencia_exportar.html', ctx)

    # Modo verificación rápida
    if request.GET.get('check') == '1':
        from django.http import JsonResponse
        count = AsistenciaSesion.objects.filter(
            curso=curso, fecha__range=(fecha_desde, fecha_hasta)
        ).count()
        return JsonResponse({'tiene_datos': count > 0, 'total': count})

    # Parsear mes del rango
    try:
        parts = fecha_desde.split('-')
        year, month = int(parts[0]), int(parts[1])
    except (ValueError, IndexError):
        ctx['error'] = 'Formato de fecha inválido.'
        return render(request, 'director/asistencia_exportar.html', ctx)

    # Días hábiles (L-V)
    _, last_day = calendar.monthrange(year, month)
    dias_habiles = []
    for d in range(1, last_day + 1):
        dt = date(year, month, d)
        if dt.weekday() < 5:
            dias_habiles.append({
                'fecha': dt,
                'dia_nombre': _DIAS_ES[dt.weekday()],
                'dia_num': str(d).zfill(2),
            })

    fecha_ini = date(year, month, 1)
    fecha_fin = date(year, month, last_day)

    # Estudiantes activos del curso
    estudiantes_qs = (
        Estudiante.objects
        .filter(curso=curso, activo=True)
        .order_by('apellido_paterno', 'apellido_materno', 'nombre')
    )

    # Asistencias del mes → mapa: estudiante_id → {fecha → letra}
    asistencias = (
        Asistencia.objects
        .filter(sesion__curso=curso, sesion__fecha__range=(fecha_ini, fecha_fin))
        .select_related('sesion')
    )
    mapa = {}
    for a in asistencias:
        mapa.setdefault(a.estudiante_id, {})[a.sesion.fecha] = _ESTADO_LETRA.get(a.estado, '')

    # Construir datos por estudiante
    estudiantes_data = []
    for i, est in enumerate(estudiantes_qs, 1):
        est_mapa = mapa.get(est.id, {})
        celdas = []
        resumen = {'presentes': 0, 'faltas': 0, 'atrasos': 0, 'licencias': 0}
        for dh in dias_habiles:
            letra = est_mapa.get(dh['fecha'], '')
            celdas.append(letra)
            if letra == 'P':
                resumen['presentes'] += 1
            elif letra == 'F':
                resumen['faltas'] += 1
            elif letra == 'A':
                resumen['atrasos'] += 1
            elif letra == 'L':
                resumen['licencias'] += 1

        estudiantes_data.append({
            'numero': i,
            'nombre': f"{est.apellido_paterno} {est.apellido_materno}".strip() + f", {est.nombre}",
            'celdas': celdas,
            'resumen': resumen,
        })

    ctx.update({
        'curso': curso,
        'mes_display': f"{_MESES_ES[month]} {year}",
        'dias_habiles': dias_habiles,
        'estudiantes': estudiantes_data,
    })
    return render(request, 'director/asistencia_exportar.html', ctx)


def director_asistencia_exportar_excel_view(request):
    import calendar
    import io
    from datetime import date
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from backend.apps.academics.models import Curso
    from backend.apps.attendance.models import AsistenciaSesion, Asistencia
    from backend.apps.students.models import Estudiante

    # ── Autenticación ────────────────────────────────────────────
    user = _autenticar_por_token(request)
    if not user or not user.tipo_usuario or user.tipo_usuario.nombre not in ('Director', 'Regente'):
        return JsonResponse({'errores': 'No autorizado.'}, status=403)

    curso_id   = request.GET.get('curso_id', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    if not curso_id or not fecha_desde:
        return JsonResponse({'errores': 'Faltan parámetros.'}, status=400)

    try:
        curso = Curso.objects.get(pk=int(curso_id))
    except (Curso.DoesNotExist, ValueError):
        return JsonResponse({'errores': 'Curso no encontrado.'}, status=404)

    try:
        parts = fecha_desde.split('-')
        year, month = int(parts[0]), int(parts[1])
    except (ValueError, IndexError):
        return JsonResponse({'errores': 'Formato de fecha inválido.'}, status=400)

    # Días hábiles del mes
    _, last_day = calendar.monthrange(year, month)
    dias_habiles = []
    for d in range(1, last_day + 1):
        dt = date(year, month, d)
        if dt.weekday() < 5:
            dias_habiles.append({'fecha': dt, 'dia_nombre': _DIAS_ES[dt.weekday()], 'dia_num': d})

    fecha_ini = date(year, month, 1)
    fecha_fin = date(year, month, last_day)

    estudiantes_qs = (
        Estudiante.objects
        .filter(curso=curso, activo=True)
        .order_by('apellido_paterno', 'apellido_materno', 'nombre')
    )

    asistencias = (
        Asistencia.objects
        .filter(sesion__curso=curso, sesion__fecha__range=(fecha_ini, fecha_fin))
        .select_related('sesion')
    )
    mapa = {}
    for a in asistencias:
        mapa.setdefault(a.estudiante_id, {})[a.sesion.fecha] = _ESTADO_LETRA.get(a.estado, '')

    # ── Construir Excel ──────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"Asistencia {_MESES_ES[month]}"

    # Estilos
    COLOR_HEADER  = '1E293B'
    COLOR_FALTA   = 'DC2626'
    COLOR_ATRASO  = 'D97706'
    COLOR_LIC     = '2563EB'
    COLOR_PRES    = '16A34A'
    COLOR_F_BG    = 'FECACA'
    COLOR_A_BG    = 'FDE68A'
    COLOR_L_BG    = 'BFDBFE'
    COLOR_P_BG    = 'BBF7D0'

    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(color='FFFFFF', bold=True, size=10):
        return Font(name='Calibri', bold=bold, color=color, size=size)

    def cell_font(color='1A1A2E', bold=False, size=10):
        return Font(name='Calibri', bold=bold, color=color, size=size)

    def fill(hex_color):
        return PatternFill(fill_type='solid', fgColor=hex_color)

    center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    left   = Alignment(horizontal='left',   vertical='center')
    rotate = Alignment(horizontal='center', vertical='bottom', text_rotation=90)

    # ── Fila 1: título escuela ───────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(dias_habiles) + 4)
    c = ws.cell(row=1, column=1, value='Unidad Educativa "Francia A" — Sucre, Chuquisaca')
    c.font = Font(name='Calibri', bold=True, size=13, color=COLOR_HEADER)
    c.alignment = center
    ws.row_dimensions[1].height = 22

    # ── Fila 2: subtítulo planilla ───────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3 + len(dias_habiles) + 4)
    mes_str = f"{_MESES_ES[month]} {year}"
    c = ws.cell(row=2, column=1, value=f"Planilla de Asistencia — Curso: {curso} — {mes_str}")
    c.font = Font(name='Calibri', bold=True, size=11, color='334155')
    c.alignment = center
    ws.row_dimensions[2].height = 18

    # ── Fila 3: cabecera de columnas ─────────────────────────────
    ROW_HDR = 3
    ws.cell(row=ROW_HDR, column=1, value='N°').font        = hdr_font()
    ws.cell(row=ROW_HDR, column=1).fill                    = fill(COLOR_HEADER)
    ws.cell(row=ROW_HDR, column=1).alignment               = center
    ws.cell(row=ROW_HDR, column=1).border                  = border

    ws.cell(row=ROW_HDR, column=2, value='Apellidos y Nombre').font = hdr_font()
    ws.cell(row=ROW_HDR, column=2).fill                    = fill(COLOR_HEADER)
    ws.cell(row=ROW_HDR, column=2).alignment               = left
    ws.cell(row=ROW_HDR, column=2).border                  = border

    # Días hábiles
    for col_idx, dh in enumerate(dias_habiles, start=3):
        lbl = f"{dh['dia_nombre'][:3]}\n{str(dh['dia_num']).zfill(2)}"
        c = ws.cell(row=ROW_HDR, column=col_idx, value=lbl)
        c.font      = hdr_font(size=8)
        c.fill      = fill(COLOR_HEADER)
        c.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        c.border    = border

    # Columnas de resumen
    resumen_cols = [
        ('Faltas',     COLOR_FALTA),
        ('Atrasos',    COLOR_ATRASO),
        ('Licencias',  COLOR_LIC),
        ('Asistencia', COLOR_PRES),
    ]
    col_res_start = 3 + len(dias_habiles)
    for i, (lbl, color) in enumerate(resumen_cols):
        c = ws.cell(row=ROW_HDR, column=col_res_start + i, value=lbl)
        c.font      = hdr_font(size=8)
        c.fill      = fill(color)
        c.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        c.border    = border

    ws.row_dimensions[ROW_HDR].height = 46

    # ── Filas de datos ───────────────────────────────────────────
    LETRAS_COLOR = {'P': COLOR_PRES, 'F': COLOR_FALTA, 'A': COLOR_ATRASO, 'L': COLOR_LIC}
    LETRAS_BG    = {'P': COLOR_P_BG, 'F': COLOR_F_BG,  'A': COLOR_A_BG,   'L': COLOR_L_BG}

    for row_i, est in enumerate(estudiantes_qs, start=1):
        row = ROW_HDR + row_i
        est_mapa = mapa.get(est.id, {})
        resumen  = {'P': 0, 'F': 0, 'A': 0, 'L': 0}
        bg_row   = 'F8FAFC' if row_i % 2 == 0 else 'FFFFFF'

        # N°
        c = ws.cell(row=row, column=1, value=row_i)
        c.font = cell_font(size=9, color='94A3B8'); c.alignment = center; c.border = border
        c.fill = fill(bg_row)

        # Nombre
        c = ws.cell(row=row, column=2, value=f"{est.apellido_paterno} {est.apellido_materno}".strip() + f", {est.nombre}")
        c.font = cell_font(); c.alignment = left; c.border = border; c.fill = fill(bg_row)

        # Celdas de días
        for col_idx, dh in enumerate(dias_habiles, start=3):
            letra = est_mapa.get(dh['fecha'], '')
            c = ws.cell(row=row, column=col_idx, value=letra)
            c.alignment = center; c.border = border
            if letra in LETRAS_COLOR:
                c.font = Font(name='Calibri', bold=True, color=LETRAS_COLOR[letra], size=10)
                c.fill = fill(LETRAS_BG[letra])
                resumen[letra] += 1
            else:
                c.font = cell_font(color='CBD5E1'); c.fill = fill(bg_row)

        # Resumen
        RES_BG = [COLOR_F_BG, COLOR_A_BG, COLOR_L_BG, COLOR_P_BG]
        RES_FG = [COLOR_FALTA, COLOR_ATRASO, COLOR_LIC, COLOR_PRES]
        RES_KEYS = ['F', 'A', 'L', 'P']
        for i, key in enumerate(RES_KEYS):
            val = resumen[key]
            c = ws.cell(row=row, column=col_res_start + i, value=val)
            c.font      = Font(name='Calibri', bold=True, color=RES_FG[i], size=10)
            c.fill      = fill(RES_BG[i])
            c.alignment = center
            c.border    = border

        ws.row_dimensions[row].height = 16

    # ── Anchos de columna ────────────────────────────────────────
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    for col_idx in range(3, 3 + len(dias_habiles)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5
    for i in range(4):
        ws.column_dimensions[get_column_letter(col_res_start + i)].width = 9

    # Congelar encabezado y columnas N°/Nombre
    ws.freeze_panes = 'C4'

    # ── Respuesta HTTP ───────────────────────────────────────────
    nombre_archivo = f"Asistencia_{curso}_{_MESES_ES[month]}_{year}.xlsx".replace(' ', '_')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


def director_estadisticas_view(request):
    return render(request, 'director/estadisticas.html')


def director_actividad_view(request):
    return render(request, 'director/actividad.html')


def director_control_diario_view(request):
    return render(request, 'director/control_diario.html')


def director_comunicados_view(request):
    return render(request, 'director/comunicados.html')


def director_academico_view(request):
    return render(request, 'director/academico.html')


def director_mi_perfil_view(request):
    return render(request, 'director/mi_perfil.html')


def profesor_view(request):
    return render(request, 'profesor/notas.html', {'active_nav': 'notas'})


def profesor_citaciones_view(request):
    return render(request, 'profesor/citaciones.html', {'active_nav': 'citaciones'})


def profesor_plan_view(request):
    return render(request, 'profesor/plan_trabajo.html', {'active_nav': 'plan'})


def profesor_cuenta_view(request):
    return render(request, 'profesor/cuenta.html', {'active_nav': 'cuenta'})


def director_comparar_nombres_view(request):
    return render(request, 'director/comparar_nombres.html')


def profesor_carga_calificaciones_view(request):
    return render(request, 'profesor/carga_calificaciones.html')

