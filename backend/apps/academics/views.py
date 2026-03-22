from datetime import date, timedelta
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db.models import Count
from django.http import HttpResponse
from rest_framework.generics import ListAPIView
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status

from backend.apps.users.permissions import IsDirectorOrRegente, IsProfesor, IsDirector
from .models import Curso, Materia, ProfesorCurso, PlanDeTrabajo, ProfesorPlan
from .serializers import CursoSerializer, MateriaSerializer, AsignacionSerializer, ProfesorPlanSerializer, DirectorPlanSerializer, ProfesorAsignacionSerializer
from .services.planilla_validator import validar_estructura, validar_pertenencia, extraer_notas, validar_estudiantes


class CursoListView(ListAPIView):
    """
    Endpoint que permite obtener el listado de cursos
    (aulas) registrados en la institución, con conteo de estudiantes.
    """

    serializer_class = CursoSerializer
    permission_classes = (IsAuthenticated, IsDirectorOrRegente)

    def get_queryset(self):
        return (
            Curso.objects
            .annotate(estudiantes_count=Count('estudiante'))
            .order_by('grado', 'paralelo')
        )


class ProfesorCursosView(APIView):
    """
    GET /api/academics/profesor/cursos/

    Retorna los cursos asignados al profesor autenticado.
    Permiso: solo Profesor.
    """

    permission_classes = [IsAuthenticated, IsProfesor]

    def get(self, request):
        cursos = (
            ProfesorCurso.objects
            .filter(profesor=request.user)
            .select_related('curso')
            .values('curso__id', 'curso__grado', 'curso__paralelo')
            .distinct()
            .order_by('curso__grado', 'curso__paralelo')
        )
        data = [
            {"id": c['curso__id'], "grado": c['curso__grado'], "paralelo": c['curso__paralelo']}
            for c in cursos
        ]
        return Response(data)


class MateriaListCreateView(APIView):
    """
    GET  /api/academics/materias/  — lista todas las materias (IsDirector)
    POST /api/academics/materias/  — crea una nueva materia (IsDirector)
    """

    permission_classes = [IsAuthenticated, IsDirector]

    def get(self, request):
        materias = Materia.objects.all().order_by('nombre')
        return Response(MateriaSerializer(materias, many=True).data)

    def post(self, request):
        serializer = MateriaSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        nombre = serializer.validated_data['nombre'].strip()
        if Materia.objects.filter(nombre__iexact=nombre).exists():
            return Response({"errores": "Ya existe una materia con ese nombre."}, status=status.HTTP_400_BAD_REQUEST)
        materia = serializer.save(nombre=nombre)
        return Response(MateriaSerializer(materia).data, status=status.HTTP_201_CREATED)


class MateriaDetailView(APIView):
    """
    DELETE /api/academics/materias/{id}/  — elimina una materia (IsDirector)
    """

    permission_classes = [IsAuthenticated, IsDirector]

    def delete(self, request, materia_id):
        try:
            materia = Materia.objects.get(pk=materia_id)
        except Materia.DoesNotExist:
            return Response({"errores": "Materia no encontrada."}, status=status.HTTP_404_NOT_FOUND)

        if ProfesorCurso.objects.filter(materia=materia).exists():
            return Response(
                {"errores": "No se puede eliminar: la materia tiene asignaciones activas."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        materia.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MateriasXCursoView(APIView):
    """
    GET /api/academics/cursos/{curso_id}/materias/

    Devuelve las materias asignadas a un curso con el nombre del profesor.
    Permiso: Director o Regente.
    """
    permission_classes = [IsAuthenticated, IsDirectorOrRegente]

    def get(self, request, curso_id):
        from django.shortcuts import get_object_or_404
        from .models import Curso
        get_object_or_404(Curso, pk=curso_id)

        qs = (
            ProfesorCurso.objects
            .filter(curso_id=curso_id)
            .select_related('materia', 'profesor')
            .order_by('materia__nombre')
        )
        data = [
            {
                'materia_id':   pc.materia.id,
                'materia':      pc.materia.nombre,
                'profesor_id':  pc.profesor.id,
                'profesor':     (f"{pc.profesor.first_name} {pc.profesor.last_name}".strip()
                                 or pc.profesor.username),
            }
            for pc in qs
        ]
        return Response(data)


class AsignacionListCreateView(APIView):
    """
    GET  /api/academics/asignaciones/  — lista todas las asignaciones Profesor-Curso-Materia (IsDirector)
    POST /api/academics/asignaciones/  — crea una nueva asignación (IsDirector)
    """

    permission_classes = [IsAuthenticated, IsDirector]

    def get(self, request):
        qs = (
            ProfesorCurso.objects
            .select_related('profesor', 'curso', 'materia')
            .order_by('curso__grado', 'curso__paralelo', 'materia__nombre')
        )
        return Response(AsignacionSerializer(qs, many=True).data)

    def post(self, request):
        serializer = AsignacionSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        profesor = serializer.validated_data['profesor']
        curso    = serializer.validated_data['curso']
        materia  = serializer.validated_data['materia']

        tipo = getattr(profesor.tipo_usuario, 'nombre', None)
        if tipo != 'Profesor':
            return Response({"errores": "El usuario seleccionado no es un Profesor."}, status=status.HTTP_400_BAD_REQUEST)

        if ProfesorCurso.objects.filter(profesor=profesor, curso=curso, materia=materia).exists():
            return Response({"errores": "Esta asignación ya existe."}, status=status.HTTP_400_BAD_REQUEST)

        asignacion = serializer.save()
        return Response(AsignacionSerializer(asignacion).data, status=status.HTTP_201_CREATED)


class ProfesorMisAsignacionesView(APIView):
    """
    GET /api/academics/profesor/mis-asignaciones/
    Retorna todas las asignaciones (ProfesorCurso) del profesor autenticado.
    """
    permission_classes = [IsAuthenticated, IsProfesor]

    def get(self, request):
        qs = (
            ProfesorCurso.objects
            .filter(profesor=request.user)
            .select_related('materia', 'curso')
            .order_by('materia__nombre', 'curso__grado', 'curso__paralelo')
        )
        return Response(ProfesorAsignacionSerializer(qs, many=True).data)


class ProfesorPlanListCreateView(APIView):
    """
    GET  /api/academics/profesor/planes/?mes=N  — planes del mes (1-12) del profesor autenticado
    POST /api/academics/profesor/planes/        — registra un nuevo plan (máx. 4 por mes)
    """

    permission_classes = [IsAuthenticated, IsProfesor]

    def get(self, request):
        mes = request.query_params.get('mes')
        qs  = ProfesorPlan.objects.filter(
            profesor_curso__profesor=request.user
        ).select_related('plan', 'profesor_curso__materia', 'profesor_curso__curso')
        if mes:
            try:
                mes = int(mes)
            except (ValueError, TypeError):
                return Response({'errores': 'El parámetro mes debe ser un número entre 1 y 12.'}, status=status.HTTP_400_BAD_REQUEST)
            qs = qs.filter(mes=mes)
        return Response(ProfesorPlanSerializer(qs.order_by('profesor_curso__materia__nombre', 'fecha_creacion'), many=True).data)

    def post(self, request):
        try:
            mes = int(request.data.get('mes', 0))
            if not (1 <= mes <= 12):
                raise ValueError
        except (ValueError, TypeError):
            return Response({'errores': 'El mes debe ser un número entre 1 y 12.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            semana = int(request.data.get('semana', 0))
            if not (1 <= semana <= 4):
                raise ValueError
        except (ValueError, TypeError):
            return Response({'errores': 'La semana debe ser un número entre 1 y 4.'}, status=status.HTTP_400_BAD_REQUEST)

        descripcion = (request.data.get('descripcion') or '').strip()
        if not descripcion:
            return Response({'errores': 'La descripción es requerida.'}, status=status.HTTP_400_BAD_REQUEST)
        if len(descripcion) > 500:
            return Response({'errores': 'La descripción no puede superar los 500 caracteres.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            profesor_curso_id = int(request.data.get('profesor_curso_id', 0))
            profesor_curso = ProfesorCurso.objects.select_related('materia', 'curso').get(
                pk=profesor_curso_id, profesor=request.user
            )
        except (ValueError, TypeError, ProfesorCurso.DoesNotExist):
            return Response({'errores': 'Asignación no válida o no te pertenece.'}, status=status.HTTP_400_BAD_REQUEST)

        if ProfesorPlan.objects.filter(profesor_curso=profesor_curso, mes=mes).count() >= 4:
            return Response(
                {'errores': f'Ya tienes 4 planes registrados para {profesor_curso.materia.nombre} en este mes. No se pueden agregar más.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        año = date.today().year
        primer_dia       = date(año, mes, 1)
        dias_hasta_lunes = (7 - primer_dia.weekday()) % 7
        primer_lunes     = primer_dia + timedelta(days=dias_hasta_lunes)
        fecha_inicio     = primer_lunes + timedelta(weeks=semana - 1)
        fecha_fin        = fecha_inicio + timedelta(days=6)

        if ProfesorPlan.objects.filter(
            profesor_curso=profesor_curso, mes=mes, plan__fecha_inicio=fecha_inicio
        ).exists():
            return Response(
                {'errores': f'Ya tienes un plan para {profesor_curso.materia.nombre} en la Semana {semana} de este mes.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        plan = PlanDeTrabajo.objects.create(descripcion=descripcion, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
        pp   = ProfesorPlan.objects.create(profesor_curso=profesor_curso, plan=plan, mes=mes)
        return Response(ProfesorPlanSerializer(pp).data, status=status.HTTP_201_CREATED)


class ProfesorPlanHistorialView(APIView):
    """
    GET /api/academics/profesor/planes/historial/
    Devuelve todos los planes de meses anteriores al actual del profesor autenticado.
    """
    permission_classes = [IsAuthenticated, IsProfesor]

    def get(self, request):
        mes_actual = date.today().month
        qs = (
            ProfesorPlan.objects
            .select_related('plan', 'profesor_curso__materia', 'profesor_curso__curso')
            .filter(profesor_curso__profesor=request.user, mes__lt=mes_actual)
            .order_by('-mes', 'profesor_curso__materia__nombre', 'plan__fecha_inicio')
        )
        return Response(ProfesorPlanSerializer(qs, many=True).data)


class ProfesorPlanDetailView(APIView):
    """
    DELETE /api/academics/profesor/planes/{plan_id}/  — elimina un plan del profesor autenticado
    """

    permission_classes = [IsAuthenticated, IsProfesor]

    def delete(self, request, plan_id):
        try:
            pp = ProfesorPlan.objects.select_related('plan').get(pk=plan_id, profesor_curso__profesor=request.user)
        except ProfesorPlan.DoesNotExist:
            return Response({'errores': 'Plan no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        plan = pp.plan
        pp.delete()
        if not ProfesorPlan.objects.filter(plan=plan).exists():
            plan.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class DirectorPlanesView(APIView):
    """
    GET /api/academics/director/planes/?mes=N[&profesor_id=X]

    Retorna los planes de trabajo de todos los profesores filtrados por mes.
    Permiso: solo Director.
    """
    permission_classes = [IsAuthenticated, IsDirector]

    def get(self, request):
        mes         = request.query_params.get('mes')
        profesor_id = request.query_params.get('profesor_id')

        qs = ProfesorPlan.objects.select_related(
            'plan',
            'profesor_curso__profesor',
            'profesor_curso__materia',
            'profesor_curso__curso',
        )

        if mes:
            try:
                mes = int(mes)
                if not (1 <= mes <= 12):
                    raise ValueError
            except (ValueError, TypeError):
                return Response(
                    {'errores': 'El parámetro mes debe ser un número entre 1 y 12.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(mes=mes)

        if profesor_id:
            qs = qs.filter(profesor_curso__profesor_id=profesor_id)

        qs = qs.order_by(
            'profesor_curso__profesor__last_name',
            'profesor_curso__profesor__first_name',
            'profesor_curso__materia__nombre',
            'plan__fecha_inicio',
        )
        return Response(DirectorPlanSerializer(qs, many=True).data)


class DirectorPlanesExportarView(APIView):
    """
    GET /api/academics/director/planes/exportar/?mes=N
    Descarga un archivo Excel con los planes de trabajo del mes indicado.
    Permiso: solo Director.
    """
    permission_classes = [IsAuthenticated, IsDirector]

    _MESES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
              'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    def get(self, request):
        mes_param = request.query_params.get('mes')
        try:
            mes = int(mes_param)
            if not (1 <= mes <= 12):
                raise ValueError
        except (ValueError, TypeError):
            return Response({'errores': 'Parámetro mes inválido (1–12).'}, status=status.HTTP_400_BAD_REQUEST)

        planes = (
            ProfesorPlan.objects
            .filter(mes=mes)
            .select_related('plan', 'profesor_curso__profesor', 'profesor_curso__materia', 'profesor_curso__curso')
        )

        profesor_id = request.query_params.get('profesor_id')
        if profesor_id:
            planes = planes.filter(profesor_curso__profesor_id=profesor_id)

        planes = planes.order_by(
            'profesor_curso__profesor__last_name',
            'profesor_curso__profesor__first_name',
            'profesor_curso__curso__grado',
            'profesor_curso__curso__paralelo',
            'profesor_curso__materia__nombre',
            'plan__fecha_inicio',
        )

        # ── Agrupar por profesor manteniendo orden ────────────────────
        from collections import OrderedDict
        por_profesor = OrderedDict()
        for pp in planes:
            pid = pp.profesor_curso.profesor.id
            if pid not in por_profesor:
                por_profesor[pid] = []
            por_profesor[pid].append(pp)

        nombre_mes = self._MESES[mes]

        # ── Estilos compartidos ───────────────────────────────────────
        fill_header  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        fill_subhead = PatternFill(start_color="2D5F8A", end_color="2D5F8A", fill_type="solid")
        font_header  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        font_sub     = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        font_body    = Font(size=10, name="Calibri")
        font_muted   = Font(size=10, color="888888", name="Calibri")
        align_c      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_l      = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
        thin_side    = Side(style="thin", color="D0D7E2")
        border       = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        headers  = ["N°", "Curso", "Materia", "Semana", "Período", "Descripción"]
        N_COLS   = len(headers)   # 6
        LAST_COL = "F"

        if not por_profesor:
            return Response(
                {'errores': f'No hay planes de trabajo registrados para {nombre_mes}.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        wb = openpyxl.Workbook()
        wb.remove(wb.active)      # eliminar hoja vacía por defecto

        for sheet_idx, (pid, pps) in enumerate(por_profesor.items(), 1):
            prof        = pps[0].profesor_curso.profesor
            nombre_prof = f"{prof.first_name} {prof.last_name}".strip() or prof.username

            sheet_name = f"{sheet_idx}. {nombre_prof}"[:31]
            ws = wb.create_sheet(title=sheet_name)

            # ── Fila 1: título del mes ────────────────────────────────
            ws.merge_cells(f"A1:{LAST_COL}1")
            tc = ws["A1"]
            tc.value     = f"Planes de Trabajo — {nombre_mes} 2026"
            tc.font      = Font(bold=True, size=14, color="1E3A5F", name="Calibri")
            tc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # ── Fila 2: nombre del profesor ───────────────────────────
            ws.merge_cells(f"A2:{LAST_COL}2")
            sc = ws["A2"]
            sc.value     = nombre_prof
            sc.font      = font_sub
            sc.fill      = fill_subhead
            sc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 22

            # ── Fila 3: cabecera de columnas ──────────────────────────
            ws.append(headers)
            for col_idx in range(1, N_COLS + 1):
                cell           = ws.cell(row=3, column=col_idx)
                cell.fill      = fill_header
                cell.font      = font_header
                cell.alignment = align_c
                cell.border    = border
            ws.row_dimensions[3].height = 22

            # ── Filas de datos ────────────────────────────────────────
            for i, pp in enumerate(pps, 1):
                d       = pp.plan.fecha_inicio.day
                semana  = 1 if d <= 7 else 2 if d <= 14 else 3 if d <= 21 else 4
                curso   = f"{pp.profesor_curso.curso.grado} \"{pp.profesor_curso.curso.paralelo}\""
                materia = pp.profesor_curso.materia.nombre
                periodo = (
                    f"{pp.plan.fecha_inicio.strftime('%d/%m/%Y')}"
                    f"  al  "
                    f"{pp.plan.fecha_fin.strftime('%d/%m/%Y')}"
                )
                fila = [i, curso, materia, f"Semana {semana}", periodo, pp.plan.descripcion]
                ws.append(fila)
                row_n = ws.max_row

                for col_idx in range(1, N_COLS + 1):
                    cell           = ws.cell(row=row_n, column=col_idx)
                    cell.font      = font_muted if col_idx == 1 else font_body
                    cell.alignment = align_l if col_idx == N_COLS else align_c
                    cell.border    = border

                lineas = max(3, len(pp.plan.descripcion) // 80 + pp.plan.descripcion.count('\n') + 1)
                ws.row_dimensions[row_n].height = lineas * 14

            # ── Anchos de columna ─────────────────────────────────────
            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 16
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 24
            ws.column_dimensions["F"].width = 68

            ws.freeze_panes = "A4"

        # ── Respuesta HTTP ────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"planes_trabajo_{nombre_mes.lower()}_2026.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class ValidarPlanillaView(APIView):
    """
    POST /api/academics/profesor/validar-planilla/

    Recibe un archivo Excel (.xlsx/.xls) y un profesor_curso_id.
    Valida que:
      1. El archivo sea una planilla Ley 070 válida.
      2. El nombre del maestro coincida con el profesor autenticado.
      3. El área/materia coincida con la materia asignada.
      4. El paralelo y año de escolaridad coincidan con el curso asignado.

    Permiso: solo Profesor.
    """

    permission_classes = [IsAuthenticated, IsProfesor]

    def post(self, request):
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'errores': 'Se requiere un archivo Excel.'}, status=status.HTTP_400_BAD_REQUEST)

        nombre = archivo.name.lower()
        if not (nombre.endswith('.xlsx') or nombre.endswith('.xls')):
            return Response({'errores': 'Solo se aceptan archivos .xlsx o .xls.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            profesor_curso_id = int(request.data.get('profesor_curso_id', 0))
        except (ValueError, TypeError):
            return Response({'errores': 'profesor_curso_id inválido.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            profesor_curso = (
                ProfesorCurso.objects
                .select_related('profesor', 'materia', 'curso')
                .get(pk=profesor_curso_id, profesor=request.user)
            )
        except ProfesorCurso.DoesNotExist:
            return Response(
                {'errores': 'Asignación no válida o no te pertenece.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        import openpyxl as _xl

        contenido = archivo.read()
        try:
            wb = _xl.load_workbook(io.BytesIO(contenido), data_only=True)
        except Exception as e:
            return Response(
                {'errores': f'No se pudo abrir el archivo: {e}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Una sola carga — se pasa el workbook abierto a todas las funciones
        resultado = validar_estructura(wb)

        if resultado['es_valido']:
            errores_pertenencia = validar_pertenencia(resultado['metadatos'], profesor_curso)
            if errores_pertenencia:
                resultado['es_valido'] = False
                resultado['errores'].extend(errores_pertenencia)
            else:
                notas = extraer_notas(wb)
                resultado['notas'] = notas

                nombres_excel = []
                for trim_data in notas['trimestres'].values():
                    datos = trim_data['saber']['datos'] or trim_data['hacer']['datos']
                    if datos:
                        nombres_excel = [e['nombre'] for e in datos]
                        break

                curso_nombre = f"{profesor_curso.curso.grado} \"{profesor_curso.curso.paralelo}\""
                val_est = validar_estudiantes(nombres_excel, profesor_curso.curso_id)
                val_est['curso_verificado'] = curso_nombre
                resultado['estudiantes'] = val_est
                if not val_est['es_valido']:
                    resultado['es_valido'] = False
                    resultado['errores'].append(
                        f"{len(val_est['no_encontrados'])} estudiante(s) del Excel "
                        f"no se encontraron en el curso {curso_nombre}."
                    )

        return Response(resultado)


class AsignacionDetailView(APIView):
    """
    DELETE /api/academics/asignaciones/{id}/  — elimina una asignación (IsDirector)
    """

    permission_classes = [IsAuthenticated, IsDirector]

    def delete(self, request, asignacion_id):
        try:
            asignacion = ProfesorCurso.objects.get(pk=asignacion_id)
        except ProfesorCurso.DoesNotExist:
            return Response({"errores": "Asignación no encontrada."}, status=status.HTTP_404_NOT_FOUND)
        asignacion.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
