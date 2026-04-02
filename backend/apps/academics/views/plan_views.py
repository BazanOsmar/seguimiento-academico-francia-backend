import io
from collections import OrderedDict
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status

from backend.core.permissions import IsDirector, IsProfesor
from ..models import ProfesorCurso, PlanDeTrabajo, ProfesorPlan
from ..serializers import ProfesorPlanSerializer, DirectorPlanSerializer


class ProfesorPlanListCreateView(APIView):
    """
    GET  /api/academics/profesor/planes/?mes=N  — planes del mes del profesor autenticado
    POST /api/academics/profesor/planes/        — registra un nuevo plan (máx. 4 por mes)
    Permiso: Profesor.
    """
    permission_classes = [IsAuthenticated, IsProfesor]

    def get(self, request):
        mes = request.query_params.get('mes')
        qs  = ProfesorPlan.objects.filter(
            profesor_curso__profesor=request.user
        ).select_related('plan', 'profesor_curso__materia', 'profesor_curso__curso')
        if mes:
            try:
                mes = int(mes)
            except (ValueError, TypeError):
                return Response(
                    {'errores': 'El parámetro mes debe ser un número entre 1 y 12.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(mes=mes)
        return Response(
            ProfesorPlanSerializer(
                qs.order_by('profesor_curso__materia__nombre', 'fecha_creacion'),
                many=True,
            ).data
        )

    def post(self, request):
        try:
            mes = int(request.data.get('mes', 0))
            if not (1 <= mes <= 12):
                raise ValueError
        except (ValueError, TypeError):
            return Response({'errores': 'El mes debe ser un número entre 1 y 12.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            semana = int(request.data.get('semana', 0))
            if not (1 <= semana <= 4):
                raise ValueError
        except (ValueError, TypeError):
            return Response({'errores': 'La semana debe ser un número entre 1 y 4.'}, status=status.HTTP_400_BAD_REQUEST)

        descripcion = (request.data.get('descripcion') or '').strip()
        if not descripcion:
            return Response({'errores': 'La descripción es requerida.'}, status=status.HTTP_400_BAD_REQUEST)
        if len(descripcion) > 500:
            return Response({'errores': 'La descripción no puede superar los 500 caracteres.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            profesor_curso_id = int(request.data.get('profesor_curso_id', 0))
            profesor_curso = ProfesorCurso.objects.select_related('materia', 'curso').get(
                pk=profesor_curso_id, profesor=request.user
            )
        except (ValueError, TypeError, ProfesorCurso.DoesNotExist):
            return Response({'errores': 'Asignación no válida o no te pertenece.'}, status=status.HTTP_400_BAD_REQUEST)

        if ProfesorPlan.objects.filter(profesor_curso=profesor_curso, mes=mes).count() >= 4:
            return Response(
                {'errores': f'Ya tienes 4 planes registrados para {profesor_curso.materia.nombre} en este mes. No se pueden agregar más.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        año             = date.today().year
        primer_dia      = date(año, mes, 1)
        dias_hasta_lunes = (7 - primer_dia.weekday()) % 7
        primer_lunes    = primer_dia + timedelta(days=dias_hasta_lunes)
        fecha_inicio    = primer_lunes + timedelta(weeks=semana - 1)
        fecha_fin       = fecha_inicio + timedelta(days=6)

        if ProfesorPlan.objects.filter(
            profesor_curso=profesor_curso, mes=mes, plan__fecha_inicio=fecha_inicio
        ).exists():
            return Response(
                {'errores': f'Ya tienes un plan para {profesor_curso.materia.nombre} en la Semana {semana} de este mes.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        plan = PlanDeTrabajo.objects.create(descripcion=descripcion, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
        pp   = ProfesorPlan.objects.create(profesor_curso=profesor_curso, plan=plan, mes=mes)
        return Response(ProfesorPlanSerializer(pp).data, status=status.HTTP_201_CREATED)


class ProfesorPlanHistorialView(APIView):
    """
    GET /api/academics/profesor/planes/historial/
    Devuelve todos los planes de meses anteriores al actual del profesor autenticado.
    Permiso: Profesor.
    """
    permission_classes = [IsAuthenticated, IsProfesor]

    def get(self, request):
        mes_actual = date.today().month
        qs = (
            ProfesorPlan.objects
            .select_related('plan', 'profesor_curso__materia', 'profesor_curso__curso')
            .filter(profesor_curso__profesor=request.user, mes__lt=mes_actual)
            .order_by('-mes', 'profesor_curso__materia__nombre', 'plan__fecha_inicio')
        )
        return Response(ProfesorPlanSerializer(qs, many=True).data)


class ProfesorPlanDetailView(APIView):
    """
    GET    /api/academics/profesor/planes/{plan_id}/  — detalle de un plan
    DELETE /api/academics/profesor/planes/{plan_id}/  — elimina un plan
    Permiso: Profesor (solo sus propios planes).
    """
    permission_classes = [IsAuthenticated, IsProfesor]

    def get(self, request, plan_id):
        try:
            pp = ProfesorPlan.objects.select_related(
                'plan', 'profesor_curso__materia', 'profesor_curso__curso'
            ).get(pk=plan_id, profesor_curso__profesor=request.user)
        except ProfesorPlan.DoesNotExist:
            return Response({'errores': 'Plan no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(ProfesorPlanSerializer(pp).data)

    def delete(self, request, plan_id):
        try:
            pp = ProfesorPlan.objects.select_related('plan').get(
                pk=plan_id, profesor_curso__profesor=request.user
            )
        except ProfesorPlan.DoesNotExist:
            return Response({'errores': 'Plan no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        plan = pp.plan
        pp.delete()
        if not ProfesorPlan.objects.filter(plan=plan).exists():
            plan.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class DirectorPlanesView(APIView):
    """
    GET /api/academics/director/planes/?mes=N[&profesor_id=X]
    Retorna los planes de todos los profesores filtrados por mes.
    Permiso: Director.
    """
    permission_classes = [IsAuthenticated, IsDirector]

    def get(self, request):
        mes         = request.query_params.get('mes')
        profesor_id = request.query_params.get('profesor_id')

        qs = ProfesorPlan.objects.select_related(
            'plan',
            'profesor_curso__profesor',
            'profesor_curso__materia',
            'profesor_curso__curso',
        )

        if mes:
            try:
                mes = int(mes)
                if not (1 <= mes <= 12):
                    raise ValueError
            except (ValueError, TypeError):
                return Response(
                    {'errores': 'El parámetro mes debe ser un número entre 1 y 12.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(mes=mes)

        if profesor_id:
            qs = qs.filter(profesor_curso__profesor_id=profesor_id)

        qs = qs.order_by(
            'profesor_curso__profesor__last_name',
            'profesor_curso__profesor__first_name',
            'profesor_curso__materia__nombre',
            'plan__fecha_inicio',
        )
        return Response(DirectorPlanSerializer(qs, many=True).data)


class DirectorPlanesExportarView(APIView):
    """
    GET /api/academics/director/planes/exportar/?mes=N
    Descarga un Excel con los planes de trabajo del mes indicado.
    Permiso: Director.
    """
    permission_classes = [IsAuthenticated, IsDirector]

    _MESES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
              'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    def get(self, request):
        mes_param = request.query_params.get('mes')
        try:
            mes = int(mes_param)
            if not (1 <= mes <= 12):
                raise ValueError
        except (ValueError, TypeError):
            return Response({'errores': 'Parámetro mes inválido (1–12).'}, status=status.HTTP_400_BAD_REQUEST)

        planes = (
            ProfesorPlan.objects
            .filter(mes=mes)
            .select_related('plan', 'profesor_curso__profesor', 'profesor_curso__materia', 'profesor_curso__curso')
        )

        profesor_id = request.query_params.get('profesor_id')
        if profesor_id:
            planes = planes.filter(profesor_curso__profesor_id=profesor_id)

        planes = planes.order_by(
            'profesor_curso__profesor__last_name',
            'profesor_curso__profesor__first_name',
            'profesor_curso__curso__grado',
            'profesor_curso__curso__paralelo',
            'profesor_curso__materia__nombre',
            'plan__fecha_inicio',
        )

        por_profesor = OrderedDict()
        for pp in planes:
            pid = pp.profesor_curso.profesor.id
            if pid not in por_profesor:
                por_profesor[pid] = []
            por_profesor[pid].append(pp)

        nombre_mes = self._MESES[mes]

        if not por_profesor:
            return Response(
                {'errores': f'No hay planes de trabajo registrados para {nombre_mes}.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        fill_header  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        fill_subhead = PatternFill(start_color="2D5F8A", end_color="2D5F8A", fill_type="solid")
        font_header  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        font_sub     = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        font_body    = Font(size=10, name="Calibri")
        font_muted   = Font(size=10, color="888888", name="Calibri")
        align_c      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_l      = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
        thin_side    = Side(style="thin", color="D0D7E2")
        border       = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        headers  = ["N°", "Curso", "Materia", "Semana", "Período", "Descripción"]
        N_COLS   = len(headers)
        LAST_COL = "F"

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet_idx, (pid, pps) in enumerate(por_profesor.items(), 1):
            prof        = pps[0].profesor_curso.profesor
            nombre_prof = f"{prof.first_name} {prof.last_name}".strip() or prof.username
            sheet_name  = f"{sheet_idx}. {nombre_prof}"[:31]
            ws          = wb.create_sheet(title=sheet_name)

            ws.merge_cells(f"A1:{LAST_COL}1")
            tc           = ws["A1"]
            tc.value     = f"Planes de Trabajo — {nombre_mes} 2026"
            tc.font      = Font(bold=True, size=14, color="1E3A5F", name="Calibri")
            tc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            ws.merge_cells(f"A2:{LAST_COL}2")
            sc           = ws["A2"]
            sc.value     = nombre_prof
            sc.font      = font_sub
            sc.fill      = fill_subhead
            sc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 22

            ws.append(headers)
            for col_idx in range(1, N_COLS + 1):
                cell           = ws.cell(row=3, column=col_idx)
                cell.fill      = fill_header
                cell.font      = font_header
                cell.alignment = align_c
                cell.border    = border
            ws.row_dimensions[3].height = 22

            for i, pp in enumerate(pps, 1):
                d       = pp.plan.fecha_inicio.day
                semana  = 1 if d <= 7 else 2 if d <= 14 else 3 if d <= 21 else 4
                curso   = f"{pp.profesor_curso.curso.grado} \"{pp.profesor_curso.curso.paralelo}\""
                materia = pp.profesor_curso.materia.nombre
                periodo = (
                    f"{pp.plan.fecha_inicio.strftime('%d/%m/%Y')}"
                    f"  al  "
                    f"{pp.plan.fecha_fin.strftime('%d/%m/%Y')}"
                )
                ws.append([i, curso, materia, f"Semana {semana}", periodo, pp.plan.descripcion])
                row_n = ws.max_row

                for col_idx in range(1, N_COLS + 1):
                    cell           = ws.cell(row=row_n, column=col_idx)
                    cell.font      = font_muted if col_idx == 1 else font_body
                    cell.alignment = align_l if col_idx == N_COLS else align_c
                    cell.border    = border

                lineas = max(3, len(pp.plan.descripcion) // 80 + pp.plan.descripcion.count('\n') + 1)
                ws.row_dimensions[row_n].height = lineas * 14

            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 16
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 24
            ws.column_dimensions["F"].width = 68
            ws.freeze_panes = "A4"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"planes_trabajo_{nombre_mes.lower()}_2026.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
