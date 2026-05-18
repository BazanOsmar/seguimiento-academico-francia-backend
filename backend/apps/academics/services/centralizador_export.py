"""
Generador de Centralizador de Notas (.xlsx) — diseño colorido.

Estructura de cada hoja:
    Fila 1     CENTRALIZADOR DE NOTAS GESTIÓN {año}
    Fila 2     ASESOR: ____            CURSO: ____
    Fila 4-5   Cabecera de columnas (N°, NOMBRES, materias x trimestres + Prom)
    Fila 6-37  Filas de estudiantes (32 slots)
    Fila 41    Porcentaje de reprobación

La cantidad de materias por curso es dinámica.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Colores ───────────────────────────────────────────────────────────
COLOR_HEADER_DARK  = "1F3864"
COLOR_HEADER_MID   = "2E75B6"
COLOR_HEADER_LIGHT = "BDD7EE"
COLOR_PROM         = "FCE4D6"
COLOR_FOOTER       = "FFF2CC"
COLOR_WHITE        = "FFFFFF"
COLOR_GRAY_ROW     = "F2F2F2"

# ── Layout ────────────────────────────────────────────────────────────
# A = N°, B = Apellidos y Nombres (sin merge), las materias arrancan en C.
COL_INICIO_MATERIAS   = 3    # C
ROW_ESTUDIANTES_START = 6
MAX_ESTUDIANTES       = 32
ROW_ESTUDIANTES_END   = ROW_ESTUDIANTES_START + MAX_ESTUDIANTES - 1  # 37
ROW_FOOTER            = ROW_ESTUDIANTES_END + 4                      # 41


# ── Helpers de estilo ─────────────────────────────────────────────────
def _thin_border():
    side = Side(style="thin", color="AAAAAA")
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def _font(bold=False, size=9, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)


def _header_cell(cell, value, bg, txt="FFFFFF"):
    cell.value = value
    cell.font = Font(name="Arial", bold=True, color=txt, size=9)
    cell.fill = _fill(bg)
    cell.alignment = _center()
    cell.border = _thin_border()


def _col_t1(mat_idx):
    return COL_INICIO_MATERIAS + mat_idx * 4


def _col_prom(mat_idx):
    return _col_t1(mat_idx) + 3


# ── Entrada pública ───────────────────────────────────────────────────
def generar_centralizador(cursos: list, archivo_salida, gestion: int):
    wb = Workbook()
    wb.remove(wb.active)
    for curso in cursos:
        _crear_hoja(wb, curso, gestion)
    wb.save(archivo_salida)


def _crear_hoja(wb, curso: dict, gestion: int):
    ws = wb.create_sheet(title=curso["nombre_hoja"])
    materias = curso["materias"]
    last_prom_col  = _col_prom(len(materias) - 1) if materias else 2
    col_reprobadas = last_prom_col + 1
    col_sexo       = last_prom_col + 2

    _anchos(ws, materias, col_reprobadas, col_sexo)
    _cabecera(ws, curso, gestion, last_prom_col)
    _headers_materias(ws, materias, col_reprobadas, col_sexo)
    _filas_estudiantes(ws, curso["estudiantes"], materias, col_reprobadas, col_sexo)
    _footer(ws, materias, col_reprobadas, col_sexo)


# ── Anchos ────────────────────────────────────────────────────────────
def _anchos(ws, materias, col_reprobadas, col_sexo):
    ws.column_dimensions["A"].width = 4     # N°
    ws.column_dimensions["B"].width = 34    # APELLIDOS Y NOMBRES (sin merge ahora)

    for mi in range(len(materias)):
        c_t1 = _col_t1(mi)
        for c in (c_t1, c_t1 + 1, c_t1 + 2):
            ws.column_dimensions[get_column_letter(c)].width = 5
        ws.column_dimensions[get_column_letter(_col_prom(mi))].width = 6

    ws.column_dimensions[get_column_letter(col_reprobadas)].width = 5
    ws.column_dimensions[get_column_letter(col_sexo)].width = 4


# ── Cabecera (filas 1-3) ──────────────────────────────────────────────
def _cabecera(ws, curso, gestion, last_prom_col):
    ws.row_dimensions[1].height = 18

    ws.merge_cells(f"A1:{get_column_letter(last_prom_col)}1")
    c = ws["A1"]
    c.value = f"CENTRALIZADOR DE NOTAS GESTIÓN {gestion}"
    c.font = Font(name="Arial", bold=True, color=COLOR_WHITE, size=12)
    c.fill = _fill(COLOR_HEADER_DARK)
    c.alignment = _center()

    ws.row_dimensions[2].height = 14
    mid_col = max(3, last_prom_col - 5)
    ws.merge_cells(f"A2:{get_column_letter(mid_col - 1)}2")
    ws["A2"].value = f"ASESOR:  {curso.get('asesor', '')}"
    ws["A2"].font = _font(bold=True)
    ws["A2"].alignment = _left()

    ws.merge_cells(f"{get_column_letter(mid_col)}2:{get_column_letter(last_prom_col)}2")
    cv = ws[f"{get_column_letter(mid_col)}2"]
    cv.value = f"CURSO:  {curso['curso_display']}"
    cv.font = Font(name="Arial", bold=True, size=11, color=COLOR_HEADER_DARK)
    cv.alignment = _center()

    ws.row_dimensions[3].height = 6


# ── Headers de materias (filas 4-5) ───────────────────────────────────
def _headers_materias(ws, materias, col_reprobadas, col_sexo):
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 14

    ws.merge_cells("A4:A5")
    _header_cell(ws["A4"], "N°", COLOR_HEADER_MID)

    ws.merge_cells("B4:B5")
    _header_cell(ws["B4"], "APELLIDOS Y NOMBRES", COLOR_HEADER_MID)

    for mi, materia in enumerate(materias):
        c_t1 = _col_t1(mi)
        c_pr = _col_prom(mi)
        lt1 = get_column_letter(c_t1)
        lt3 = get_column_letter(c_t1 + 2)
        lpr = get_column_letter(c_pr)

        ws.merge_cells(f"{lt1}4:{lt3}4")
        _header_cell(ws[f"{lt1}4"], materia, COLOR_HEADER_MID)

        ws.merge_cells(f"{lpr}4:{lpr}5")
        _header_cell(ws[f"{lpr}4"], "Prom", COLOR_PROM, txt="000000")

        for i in range(3):
            cc = ws.cell(row=5, column=c_t1 + i)
            cc.value = i + 1
            cc.font = Font(name="Arial", bold=True, color="000000", size=8)
            cc.fill = _fill(COLOR_HEADER_LIGHT)
            cc.alignment = _center()
            cc.border = _thin_border()

    ws.merge_cells(f"{get_column_letter(col_reprobadas)}4:{get_column_letter(col_reprobadas)}5")
    _header_cell(ws.cell(row=4, column=col_reprobadas), "< 51", "C00000")

    ws.merge_cells(f"{get_column_letter(col_sexo)}4:{get_column_letter(col_sexo)}5")
    _header_cell(ws.cell(row=4, column=col_sexo), "S", COLOR_HEADER_MID)


# ── Filas de estudiantes ──────────────────────────────────────────────
def _filas_estudiantes(ws, estudiantes, materias, col_reprobadas, col_sexo):
    for off in range(MAX_ESTUDIANTES):
        row = ROW_ESTUDIANTES_START + off
        ws.row_dimensions[row].height = 14

        if off < len(estudiantes):
            _fila_alumno(ws, row, off + 1, estudiantes[off], materias, col_reprobadas, col_sexo)
        else:
            _fila_vacia(ws, row, materias, col_reprobadas, col_sexo)


def _fila_alumno(ws, row, numero, est, materias, col_reprobadas, col_sexo):
    bg = COLOR_GRAY_ROW if numero % 2 == 0 else COLOR_WHITE

    # A: N°
    c = ws.cell(row=row, column=1)
    c.value = numero
    c.font = _font()
    c.fill = _fill(bg)
    c.alignment = _center()
    c.border = _thin_border()

    # B: Nombres
    c = ws.cell(row=row, column=2)
    c.value = (est.get("nombre") or "").upper()
    c.font = _font()
    c.fill = _fill(bg)
    c.alignment = _left()
    c.border = _thin_border()

    notas = est.get("notas", {})

    for mi, materia in enumerate(materias):
        c_t1 = _col_t1(mi)
        c_pr = _col_prom(mi)
        vals = notas.get(materia, [None, None, None])

        for t in range(3):
            nota = vals[t] if t < len(vals) else None
            cell = ws.cell(row=row, column=c_t1 + t)
            cell.value = nota
            cell.font = _font()
            cell.fill = _fill(bg)
            cell.alignment = _center()
            cell.border = _thin_border()
            if nota is not None and nota < 51:
                cell.font = Font(name="Arial", size=9, color="C00000", bold=True)

        t1_l = get_column_letter(c_t1)
        t2_l = get_column_letter(c_t1 + 1)
        t3_l = get_column_letter(c_t1 + 2)
        cp = ws.cell(row=row, column=c_pr)
        cp.value = (
            f'=IF(OR({t1_l}{row}="",{t2_l}{row}="",{t3_l}{row}=""),"",'
            f'ROUND(AVERAGE({t1_l}{row}:{t3_l}{row}),0))'
        )
        cp.font = _font(bold=True)
        cp.fill = _fill(COLOR_PROM)
        cp.alignment = _center()
        cp.border = _thin_border()

    # AW: conteo notas < 51 (solo trimestres, no Prom)
    conteos = "+".join(
        f'COUNTIF({get_column_letter(_col_t1(mi))}{row}:{get_column_letter(_col_t1(mi)+2)}{row},"<51")'
        for mi in range(len(materias))
    ) or "0"
    caw = ws.cell(row=row, column=col_reprobadas)
    caw.value = f"={conteos}"
    caw.font = Font(name="Arial", size=9, color="C00000", bold=True)
    caw.fill = _fill(bg)
    caw.alignment = _center()
    caw.border = _thin_border()

    # AX: sexo
    cs = ws.cell(row=row, column=col_sexo)
    cs.value = est.get("sexo", "")
    cs.font = _font()
    cs.fill = _fill(bg)
    cs.alignment = _center()
    cs.border = _thin_border()


def _fila_vacia(ws, row, materias, col_reprobadas, col_sexo):
    last = _col_prom(len(materias) - 1) if materias else 2
    for col in range(1, max(last, col_sexo) + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _fill(COLOR_WHITE)
        c.border = _thin_border()


# ── Footer (fila 41) ──────────────────────────────────────────────────
def _footer(ws, materias, col_reprobadas, col_sexo):
    row = ROW_FOOTER
    ws.row_dimensions[row].height = 20

    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = "Porcentaje de reprobación"
    c.font = Font(name="Arial", bold=True, size=9, color=COLOR_WHITE)
    c.fill = _fill("C00000")
    c.alignment = _left()
    c.border = _thin_border()

    r_ini = ROW_ESTUDIANTES_START
    r_fin = ROW_ESTUDIANTES_END

    for mi in range(len(materias)):
        c_t1 = _col_t1(mi)
        c_pr = _col_prom(mi)

        for t in range(3):
            col_l = get_column_letter(c_t1 + t)
            cell = ws.cell(row=row, column=c_t1 + t)
            cell.value = (
                f'=IFERROR(ROUND((COUNTIF({col_l}{r_ini}:{col_l}{r_fin},"<51")*100)/'
                f'(COUNT({col_l}{r_ini}:{col_l}{r_fin})),0),"")'
            )
            cell.font = Font(name="Arial", bold=True, size=9, color="C00000")
            cell.fill = _fill(COLOR_FOOTER)
            cell.alignment = _center()
            cell.border = _thin_border()

        cpl = get_column_letter(c_pr)
        cp = ws.cell(row=row, column=c_pr)
        cp.value = (
            f'=IFERROR(ROUND((COUNTIF({cpl}{r_ini}:{cpl}{r_fin},"<51")*100)/'
            f'(COUNT({cpl}{r_ini}:{cpl}{r_fin})),0),"")'
        )
        cp.font = Font(name="Arial", bold=True, size=9, color="C00000")
        cp.fill = _fill(COLOR_FOOTER)
        cp.alignment = _center()
        cp.border = _thin_border()

    aw_l = get_column_letter(col_reprobadas)
    ax_l = get_column_letter(col_sexo)

    c_aw = ws.cell(row=row, column=col_reprobadas)
    c_aw.value = f'=COUNTIF({aw_l}{r_ini}:{aw_l}{r_fin},">0")'
    c_aw.font = Font(name="Arial", bold=True, size=9)
    c_aw.fill = _fill(COLOR_FOOTER)
    c_aw.alignment = _center()
    c_aw.border = _thin_border()

    c_ax = ws.cell(row=row, column=col_sexo)
    c_ax.value = f'=COUNTIF({ax_l}{r_ini}:{ax_l}{r_fin},"F")'
    c_ax.font = Font(name="Arial", bold=True, size=9)
    c_ax.fill = _fill(COLOR_FOOTER)
    c_ax.alignment = _center()
    c_ax.border = _thin_border()
