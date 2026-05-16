"""
Genera el archivo .xlsx de "Descargar notas" para el Director.

Una hoja por (curso, materia) del profesor. Cada hoja replica la planilla
boliviana del profesor: SER/10 + SABER/45 + HACER/40, con sus actividades
y promedios, más TOTAL, AUTOEVALUACIÓN, NOTA TRIMESTRAL y SITUACIÓN.

Modos:
  - mes = 1..12       → notas del mes indicado.
  - mes = "actualizadas" → para cada (curso, materia), última nota disponible
                          (busca el mes más reciente con notas en gestión).
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .notas_mongo_service import _get_db, obtener_notas_mes


# ── Estilo / Constantes ───────────────────────────────────────────────
_DIMS = [
    {'key': 'ser',   'label': 'SER/10',   'max': 10},
    {'key': 'saber', 'label': 'SABER/45', 'max': 45},
    {'key': 'hacer', 'label': 'HACER/40', 'max': 40},
]

C_TITLE      = "1F3864"
C_DIM        = "ED7D31"  # naranja del header SER/SABER/HACER
C_PROM       = "FFE699"  # amarillo suave
C_AUTOEVAL   = "FFD966"
C_NOTA_TRIM  = "FFE699"
C_SIT        = "9BC2E6"
C_TRIM       = "C6E0B4"
C_HEADER_BG  = "F2F2F2"
C_GRAY_ROW   = "F2F2F2"
C_WHITE      = "FFFFFF"
C_RED        = "C00000"


def _thin():
    s = Side(style="thin", color="999999")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _font(bold=False, size=9, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center")


# ── Búsqueda de último mes con notas ──────────────────────────────────
def _ultimo_mes_con_notas(profesor_id, curso_id, materia_id, gestion):
    pipeline = [
        {"$match": {
            "profesor_id": profesor_id,
            "curso_id":    curso_id,
            "materia_id":  materia_id,
            "gestion":     gestion,
        }},
        {"$group": {"_id": "$mes"}},
        {"$sort": {"_id": -1}},
        {"$limit": 1},
    ]
    docs = list(_get_db()['detalle_notas'].aggregate(pipeline))
    return docs[0]['_id'] if docs else None


def _autoeval_mes(profesor_id, curso_id, materia_id, mes, gestion):
    """Devuelve { estudiante_id: autoeval } leyendo notas_mensuales del mes."""
    cur = _get_db()['notas_mensuales'].find({
        "profesor_id":  profesor_id,
        "curso_id":     curso_id,
        "materia_id":   materia_id,
        "mes":          mes,
        "gestion":      gestion,
        "autoeval_ser": {"$ne": None},
    }, {"_id": 0, "estudiante_id": 1, "autoeval_ser": 1})
    return {d["estudiante_id"]: d["autoeval_ser"] for d in cur}


# ── Generación principal ──────────────────────────────────────────────
def generar_excel_notas(profesor_id, curso_id, mes_param, gestion) -> BytesIO | None:
    """
    Returns BytesIO con el .xlsx o None si no hay notas que exportar.
    """
    from backend.apps.academics.models import ProfesorCurso

    qs = ProfesorCurso.objects.select_related('curso', 'materia', 'profesor').filter(profesor_id=profesor_id)
    if curso_id:
        qs = qs.filter(curso_id=curso_id)
    asignaciones = list(qs.order_by('curso__grado', 'curso__paralelo', 'materia__nombre'))
    if not asignaciones:
        return None

    profesor = asignaciones[0].profesor
    prof_nombre = f"{profesor.first_name} {profesor.last_name}".strip() or profesor.username

    es_actualizadas = (mes_param == "actualizadas")
    try:
        mes_fijo = int(mes_param) if not es_actualizadas else None
    except (TypeError, ValueError):
        mes_fijo = None
        es_actualizadas = True

    wb = Workbook()
    wb.remove(wb.active)
    hojas_creadas = 0

    for asig in asignaciones:
        # Determinar mes efectivo
        if es_actualizadas:
            mes_efectivo = _ultimo_mes_con_notas(profesor_id, asig.curso_id, asig.materia_id, gestion)
        else:
            mes_efectivo = mes_fijo

        if not mes_efectivo:
            continue

        headers = obtener_notas_mes(asig.materia_id, asig.curso_id, profesor_id, mes_efectivo, gestion)
        if not headers:
            continue

        autoevals = _autoeval_mes(profesor_id, asig.curso_id, asig.materia_id, mes_efectivo, gestion)
        _crear_hoja(
            wb,
            asig=asig,
            prof_nombre=prof_nombre,
            headers=headers,
            autoevals=autoevals,
            mes_efectivo=mes_efectivo,
            gestion=gestion,
        )
        hojas_creadas += 1

    if hojas_creadas == 0:
        return None

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Construcción de la hoja ───────────────────────────────────────────
def _crear_hoja(wb, asig, prof_nombre, headers, autoevals, mes_efectivo, gestion):
    nombre_hoja = f"{asig.curso.grado.replace(' ','')}{asig.curso.paralelo}_{asig.materia.nombre[:10]}"[:31]
    ws = wb.create_sheet(title=nombre_hoja)

    # Tomar el primer trimestre que tenga datos
    trim_key = next(iter(headers))
    dim_data = headers[trim_key]   # {'ser': [...], 'saber': [...], 'hacer': [...]}

    # Recopilar columnas por dimensión (preservando orden)
    cols_por_dim = []   # [(dim_meta, [cols])]
    total_act_cols = 0
    for dim_meta in _DIMS:
        cols = dim_data.get(dim_meta['key'], [])
        cols_por_dim.append((dim_meta, cols))
        total_act_cols += len(cols)

    # Recolectar estudiantes (unión de todas las dims)
    estudiantes_map = {}
    for _, cols in cols_por_dim:
        for col in cols:
            for n in col.get('notas', []):
                if n['nro'] not in estudiantes_map:
                    estudiantes_map[n['nro']] = n['nombre']
    estudiantes = sorted(estudiantes_map.items(), key=lambda x: x[0])

    # ── Layout columnas ──────────────────────────────────────────────
    # A: N° | B: Apellidos y Nombres
    # Luego por cada dimensión: N columnas de actividad + 1 PROMEDIO
    # Después: TOTAL | AUTOEVAL | NOTA TRIM | SITUACIÓN
    col_idx = 3
    rangos_dim = []     # [(start_col, end_col_actividades, col_prom)]
    for dim_meta, cols in cols_por_dim:
        n = max(len(cols), 1)
        rangos_dim.append({
            'meta':       dim_meta,
            'cols':       cols,
            'start':      col_idx,
            'end_act':    col_idx + n - 1,
            'col_prom':   col_idx + n,
        })
        col_idx += n + 1   # +1 por la columna PROMEDIO

    col_total      = col_idx;     col_idx += 1
    col_autoeval   = col_idx;     col_idx += 1
    col_nota_trim  = col_idx;     col_idx += 1
    col_situacion  = col_idx
    last_col       = col_situacion

    # Período mostrado
    mes_nombre = ['', 'Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'][mes_efectivo]

    primera_act_col = rangos_dim[0]['start'] if rangos_dim else 3
    ultima_prom_col = rangos_dim[-1]['col_prom'] if rangos_dim else 3

    # ── FILA 1: Título institucional ─────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1)
    c.value = "UNIDAD EDUCATIVA REPÚBLICA DE FRANCIA \"A\""
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = _fill(C_TITLE)
    c.alignment = _center()
    c.border = _thin()

    # ── FILAS 2-3: Metadatos (NIVEL/ÁREA/PARALELO + CAMPO/MAESTRA/PERÍODO)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    third = max(2, (last_col) // 3)

    ws.cell(row=2, column=1).value = "NIVEL:"
    ws.cell(row=2, column=1).font = _font(bold=True, size=9)
    ws.cell(row=2, column=1).fill = _fill(C_HEADER_BG)
    ws.cell(row=2, column=1).alignment = _left()
    ws.cell(row=2, column=1).border = _thin()
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=third)
    v = ws.cell(row=2, column=2)
    v.value = "SECUNDARIO COMUNITARIO PRODUCTIVO"
    v.font = _font(size=9); v.alignment = _left(); v.border = _thin()

    ws.cell(row=2, column=third + 1).value = "ÁREA:"
    ws.cell(row=2, column=third + 1).font = _font(bold=True, size=9)
    ws.cell(row=2, column=third + 1).fill = _fill(C_HEADER_BG)
    ws.cell(row=2, column=third + 1).alignment = _left()
    ws.cell(row=2, column=third + 1).border = _thin()
    ws.merge_cells(start_row=2, start_column=third + 2, end_row=2, end_column=2 * third)
    v = ws.cell(row=2, column=third + 2)
    v.value = asig.materia.nombre
    v.font = _font(size=9); v.alignment = _left(); v.border = _thin()

    ws.cell(row=2, column=2 * third + 1).value = "PARALELO:"
    ws.cell(row=2, column=2 * third + 1).font = _font(bold=True, size=9)
    ws.cell(row=2, column=2 * third + 1).fill = _fill(C_HEADER_BG)
    ws.cell(row=2, column=2 * third + 1).alignment = _left()
    ws.cell(row=2, column=2 * third + 1).border = _thin()
    ws.merge_cells(start_row=2, start_column=2 * third + 2, end_row=2, end_column=last_col)
    v = ws.cell(row=2, column=2 * third + 2)
    v.value = asig.curso.paralelo
    v.font = _font(bold=True, size=11); v.alignment = _center(); v.border = _thin()

    # Fila 3
    ws.cell(row=3, column=1).value = "CAMPO:"
    ws.cell(row=3, column=1).font = _font(bold=True, size=9)
    ws.cell(row=3, column=1).fill = _fill(C_HEADER_BG)
    ws.cell(row=3, column=1).alignment = _left()
    ws.cell(row=3, column=1).border = _thin()
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=third)
    v = ws.cell(row=3, column=2); v.value = ""; v.font = _font(size=9); v.alignment = _left(); v.border = _thin()

    ws.cell(row=3, column=third + 1).value = "MAESTRA/O:"
    ws.cell(row=3, column=third + 1).font = _font(bold=True, size=9)
    ws.cell(row=3, column=third + 1).fill = _fill(C_HEADER_BG)
    ws.cell(row=3, column=third + 1).alignment = _left()
    ws.cell(row=3, column=third + 1).border = _thin()
    ws.merge_cells(start_row=3, start_column=third + 2, end_row=3, end_column=2 * third)
    v = ws.cell(row=3, column=third + 2)
    v.value = prof_nombre
    v.font = _font(size=9); v.alignment = _left(); v.border = _thin()

    ws.cell(row=3, column=2 * third + 1).value = "PERÍODO:"
    ws.cell(row=3, column=2 * third + 1).font = _font(bold=True, size=9)
    ws.cell(row=3, column=2 * third + 1).fill = _fill(C_HEADER_BG)
    ws.cell(row=3, column=2 * third + 1).alignment = _left()
    ws.cell(row=3, column=2 * third + 1).border = _thin()
    ws.merge_cells(start_row=3, start_column=2 * third + 2, end_row=3, end_column=last_col)
    v = ws.cell(row=3, column=2 * third + 2)
    v.value = f"{mes_nombre} {gestion}"
    v.font = _font(bold=True, size=9, color=C_TITLE)
    v.alignment = _center(); v.border = _thin()

    # ── FILA 4: Banner "EVALUACIÓN..." + headers extras merged 4-6 ────
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 75

    # N° y APELLIDOS Y NOMBRES merge fila 4-6 (3 filas)
    ws.merge_cells(start_row=4, start_column=1, end_row=6, end_column=1)
    c = ws.cell(row=4, column=1); c.value = "N°"
    c.font = _font(bold=True, size=10, color="FFFFFF"); c.fill = _fill(C_DIM)
    c.alignment = _center(); c.border = _thin()
    ws.merge_cells(start_row=4, start_column=2, end_row=6, end_column=2)
    c = ws.cell(row=4, column=2); c.value = "APELLIDOS Y NOMBRES"
    c.font = _font(bold=True, size=10, color="FFFFFF"); c.fill = _fill(C_DIM)
    c.alignment = _center(); c.border = _thin()

    # Banner EVALUACIÓN sobre dimensiones (fila 4)
    if rangos_dim:
        ws.merge_cells(start_row=4, start_column=primera_act_col, end_row=4, end_column=ultima_prom_col)
        c = ws.cell(row=4, column=primera_act_col)
        c.value = "EVALUACIÓN DEL MAESTRO AL ESTUDIANTE (95 PUNTOS)"
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = _fill(C_DIM); c.alignment = _center(); c.border = _thin()

    # Headers extras: TOTAL / AUTOEVAL / NOTA TRIM / SITUACIÓN — merge filas 4-6
    for col, label, color in [
        (col_total,     "TOTAL",       C_PROM),
        (col_autoeval,  "AUTOEVAL.",   C_AUTOEVAL),
        (col_nota_trim, "NOTA TRIM.",  C_NOTA_TRIM),
        (col_situacion, "SITUACIÓN",   C_SIT),
    ]:
        ws.merge_cells(start_row=4, start_column=col, end_row=6, end_column=col)
        c = ws.cell(row=4, column=col)
        c.value = label
        c.font = _font(bold=True, size=10)
        c.fill = _fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin()

    # ── FILA 5: Headers de dimensiones (SER/SABER/HACER) ─────────────
    for rg in rangos_dim:
        ws.merge_cells(start_row=5, start_column=rg['start'], end_row=5, end_column=rg['col_prom'])
        c = ws.cell(row=5, column=rg['start'])
        c.value = rg['meta']['label']
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = _fill(C_DIM); c.alignment = _center(); c.border = _thin()

    # ── FILA 6: Títulos de actividad rotados + PROMEDIO ──────────────
    for rg in rangos_dim:
        for i, col_data in enumerate(rg['cols']):
            c = ws.cell(row=6, column=rg['start'] + i)
            c.value = col_data.get('titulo', f"Act {i+1}")
            c.font = _font(bold=True, size=7)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
            c.fill = _fill(C_HEADER_BG)
            c.border = _thin()
        if not rg['cols']:
            c = ws.cell(row=6, column=rg['start'])
            c.value = "—"
            c.font = _font(size=8); c.alignment = _center(); c.border = _thin()
        cp = ws.cell(row=6, column=rg['col_prom'])
        cp.value = "PROMEDIO"
        cp.font = _font(bold=True, size=8)
        cp.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
        cp.fill = _fill(C_PROM)
        cp.border = _thin()

    # ── Anchos ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    for rg in rangos_dim:
        for i in range(rg['end_act'] - rg['start'] + 1):
            ws.column_dimensions[get_column_letter(rg['start'] + i)].width = 3.5
        ws.column_dimensions[get_column_letter(rg['col_prom'])].width = 7
    ws.column_dimensions[get_column_letter(col_total)].width      = 7
    ws.column_dimensions[get_column_letter(col_autoeval)].width   = 9
    ws.column_dimensions[get_column_letter(col_nota_trim)].width  = 8
    ws.column_dimensions[get_column_letter(col_situacion)].width  = 11

    # ── Datos de estudiantes ────────────────────────────────────────
    row = 7
    for idx, (nro, nombre) in enumerate(estudiantes, start=1):
        bg = C_GRAY_ROW if idx % 2 == 0 else C_WHITE

        c = ws.cell(row=row, column=1)
        c.value = idx
        c.font = _font(size=9, bold=True)
        c.fill = _fill(bg)
        c.alignment = _center()
        c.border = _thin()

        c = ws.cell(row=row, column=2)
        c.value = nombre
        c.font = _font(size=9)
        c.fill = _fill(bg)
        c.alignment = _left()
        c.border = _thin()

        promedios_dim = {}
        for rg in rangos_dim:
            dim_key = rg['meta']['key']
            dim_max = rg['meta']['max']
            suma, contador = 0.0, 0
            for i, col_data in enumerate(rg['cols']):
                nota = None
                for n in col_data.get('notas', []):
                    if n['nro'] == nro:
                        nota = n['nota']
                        break
                cell = ws.cell(row=row, column=rg['start'] + i)
                cell.value = nota
                cell.font = _font(size=9)
                cell.fill = _fill(bg)
                cell.alignment = _center()
                cell.border = _thin()
                if nota is not None:
                    suma += float(nota)
                    contador += 1
            promedio = round(suma / contador, 1) if contador else None
            if promedio is not None and promedio > dim_max:
                promedio = dim_max
            promedios_dim[dim_key] = promedio
            cp = ws.cell(row=row, column=rg['col_prom'])
            cp.value = promedio
            cp.font = _font(size=9, bold=True)
            cp.fill = _fill(C_PROM)
            cp.alignment = _center()
            cp.border = _thin()

        total = sum(p for p in promedios_dim.values() if p is not None)
        c = ws.cell(row=row, column=col_total)
        c.value = round(total, 1) if any(promedios_dim.values()) else None
        c.font = _font(size=9, bold=True)
        c.fill = _fill(C_PROM)
        c.alignment = _center()
        c.border = _thin()

        autoeval = autoevals.get(nro)
        c = ws.cell(row=row, column=col_autoeval)
        c.value = autoeval
        c.font = _font(size=9)
        c.fill = _fill(C_AUTOEVAL)
        c.alignment = _center()
        c.border = _thin()

        nota_trim = total + (autoeval if autoeval else 0)
        c = ws.cell(row=row, column=col_nota_trim)
        c.value = round(nota_trim, 1) if total or autoeval else None
        c.font = _font(size=9, bold=True, color=(C_RED if nota_trim < 51 else "000000"))
        c.fill = _fill(C_NOTA_TRIM)
        c.alignment = _center()
        c.border = _thin()

        c = ws.cell(row=row, column=col_situacion)
        if total or autoeval:
            c.value = "Aprobado" if nota_trim >= 51 else "Reprobado"
            c.font = _font(size=9, bold=True, color=("008000" if nota_trim >= 51 else C_RED))
        c.fill = _fill(C_SIT)
        c.alignment = _center()
        c.border = _thin()

        row += 1
