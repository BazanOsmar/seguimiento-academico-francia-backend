"""
Validador de planillas Excel del Ministerio de Educación (Ley 070).

Verifica que el archivo sea una planilla de notas válida y comprueba
que los metadatos correspondan al ProfesorCurso indicado.
"""

import re
import unicodedata
import io

import openpyxl
from django.utils import timezone


# ── Mensaje genérico de estructura inválida ────────────────────────────────────
_MENSAJE_NO_OFICIAL = (
    "Esta plantilla no es la oficial del Ministerio de Educación. "
    "Verifica que estés cargando la planilla correcta."
)

# ── Hojas y estructura esperadas ──────────────────────────────────────────────
HOJAS_OBLIGATORIAS = ['CARATULA', 'FILIACION', '1TRIM', '2TRIM', '3TRIM', 'BOLETIN']

ESTRUCTURA_TRIMESTRE = {
    'N8': 'SER/10',
    'S8': 'SABER/45',
    'AD8': 'HACER/40',
    'AO8': 'TOTAL',
    'AQ7': 'NOTA TRIMESTRAL',
    'AU7': 'SITUACIÓN TRIMESTRAL',
    'AP7': 'AUTOEVALUACIÓN (5 PUNTOS)',
}

TRIMESTRES = ['1TRIM', '2TRIM', '3TRIM']

META_CARATULA = {
    'maestro':         'F12',
    'campo':           'B20',
    'area':            'E20',
    'año_escolaridad': 'H20',
    'paralelo':        'J20',
    'director':        'F10',
    'gestion':         'F14',
    'nivel':           'F16',
    'unidad_educativa':'F8',
}

META_BOLETIN = {
    'maestro':         'J9',
    'campo':           'J7',
    'area':            'Z5',
    'año_escolaridad': 'AB7',
    'paralelo':        'AB9',
    'director':        'J11',
}

META_TRIMESTRE = {
    'maestro':         'AA5',
    'campo':           'D5',
    'area':            'AA3',
    'año_escolaridad': 'AR1',
    'paralelo':        'AR3',
}


def _leer_celda(ws, cell_ref):
    val = ws[cell_ref].value
    if val is None:
        return None
    val_str = str(val).strip()
    return val_str if val_str else None


def _normalizar(texto):
    """Lowercase, sin tildes/diacríticos, solo alfanumérico y espacios."""
    if not texto:
        return ''
    texto = unicodedata.normalize('NFKD', str(texto))
    texto = texto.encode('ascii', 'ignore').decode('ascii')
    texto = texto.lower()
    texto = re.sub(r'[^a-z0-9 ]', ' ', texto)
    return ' '.join(texto.split())


def _base_ordinal(s):
    """Normaliza ordinales: '1ro', '1°', 'Primero' → '1'."""
    _ORDINALES = {
        'primero': '1', 'primera': '1',
        'segundo': '2', 'segunda': '2',
        'tercero': '3', 'tercera': '3',
        'cuarto':  '4', 'cuarta':  '4',
        'quinto':  '5', 'quinta':  '5',
        'sexto':   '6', 'sexta':   '6',
    }
    for palabra, num in _ORDINALES.items():
        s = re.sub(rf'\b{palabra}\b', num, s)
    s = re.sub(r'[°º]', '', s)
    s = re.sub(r'\b(\d+)(ro|do|er|to|vo|mo|no)\b', r'\1', s)
    return s.strip()


# ── Validación de estructura ───────────────────────────────────────────────────

def validar_estructura(archivo):
    """
    Valida que el archivo sea una planilla Ley 070 oficial.
    Retorna: { es_valido, mensaje, advertencias[], metadatos{} }
    En cuanto falla cualquier verificación estructural, retorna el mensaje
    genérico y detiene la validación.
    """
    resultado = {
        'es_valido': True,
        'mensaje':   None,
        'advertencias': [],
        'metadatos': {},
    }

    if isinstance(archivo, openpyxl.Workbook):
        wb = archivo
    else:
        if isinstance(archivo, (bytes, bytearray)):
            archivo = io.BytesIO(archivo)
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
        except Exception:
            resultado['es_valido'] = False
            resultado['mensaje']   = _MENSAJE_NO_OFICIAL
            return resultado

    # 1. Hojas obligatorias
    hojas = wb.sheetnames
    for hoja in HOJAS_OBLIGATORIAS:
        if hoja not in hojas:
            resultado['es_valido'] = False
            resultado['mensaje']   = _MENSAJE_NO_OFICIAL
            return resultado

    # 2. Estructura de cada hoja de trimestre
    for trim in TRIMESTRES:
        ws = wb[trim]
        for cell_ref, valor_esperado in ESTRUCTURA_TRIMESTRE.items():
            if _leer_celda(ws, cell_ref) != valor_esperado:
                resultado['es_valido'] = False
                resultado['mensaje']   = _MENSAJE_NO_OFICIAL
                return resultado

    # 3. Extraer metadatos (CARATULA → BOLETIN → 1TRIM como fallback)
    ws_car  = wb['CARATULA']
    ws_bol  = wb['BOLETIN']
    ws_trim = wb['1TRIM']
    metadatos_extraidos = {}
    for campo, cell_caratula in META_CARATULA.items():
        val = _leer_celda(ws_car, cell_caratula)
        if not val and campo in META_BOLETIN:
            val = _leer_celda(ws_bol, META_BOLETIN[campo])
        if not val and campo in META_TRIMESTRE:
            val = _leer_celda(ws_trim, META_TRIMESTRE[campo])
        metadatos_extraidos[campo] = val
    resultado['metadatos'].update(metadatos_extraidos)

    # 4. Validar año académico (gestión)
    gestion_raw = str(metadatos_extraidos.get('gestion') or '')
    digitos_gestion = re.findall(r'\d{4}', gestion_raw)
    año_actual = timezone.now().year
    if digitos_gestion and str(año_actual) not in digitos_gestion:
        resultado['es_valido'] = False
        resultado['mensaje']   = (
            f"La planilla corresponde a la gestión {digitos_gestion[0]}, "
            f"pero el sistema está en {año_actual}."
        )
        return resultado

    # 5. Extraer estudiantes de FILIACION
    ws_fil = wb['FILIACION']
    nombres_filiacion = []
    for fila in range(9, 49):
        nombre = ws_fil.cell(row=fila, column=2).value
        if nombre and str(nombre).strip():
            nombres_filiacion.append(str(nombre).strip())
        else:
            break
    resultado['metadatos']['cantidad_estudiantes'] = len(nombres_filiacion)
    resultado['metadatos']['estudiantes']          = nombres_filiacion

    if len(nombres_filiacion) == 0:
        resultado['es_valido'] = False
        resultado['mensaje']   = "La planilla no tiene estudiantes registrados en FILIACION."
        return resultado

    # 6. Advertencia por campos críticos vacíos
    campos_criticos = ['maestro', 'paralelo', 'area', 'año_escolaridad']
    campos_vacios = [c for c in campos_criticos if not metadatos_extraidos.get(c)]
    if campos_vacios:
        resultado['advertencias'].append(
            f"Metadatos sin llenar en CARATULA: {', '.join(campos_vacios)}. "
            "Completa la carátula antes de subir la planilla."
        )

    # 7. Verificar notas por trimestre
    trims_sin_notas = []
    for trim in TRIMESTRES:
        ws = wb[trim]
        tiene_notas = False
        for col in [19, 30]:
            for fila in range(15, 55):
                val = ws.cell(row=fila, column=col).value
                if val is not None and val != '':
                    tiene_notas = True
                    break
            if tiene_notas:
                break
        resultado['metadatos'][f'{trim}_tiene_notas'] = tiene_notas
        if not tiene_notas:
            trims_sin_notas.append(trim)

    if len(trims_sin_notas) == 3:
        resultado['advertencias'].append(
            "La planilla no tiene notas en ningún trimestre. ¿Estás seguro de que es la correcta?"
        )
    else:
        for trim in trims_sin_notas:
            resultado['advertencias'].append(f"[{trim}] No contiene notas cargadas.")

    return resultado


# ── Validación de pertenencia ─────────────────────────────────────────────────

def validar_pertenencia(metadatos, profesor_curso):
    """
    Verifica en orden: nombre → grado → paralelo → área/materia.
    Retorna el primer error como string, o None si todo OK.
    """
    # 1. Nombre del maestro
    maestro_planilla = _normalizar(metadatos.get('maestro', ''))
    nombre_completo  = f"{profesor_curso.profesor.first_name} {profesor_curso.profesor.last_name}".strip()
    nombre_db        = _normalizar(nombre_completo or profesor_curso.profesor.username)

    if not maestro_planilla or (nombre_db and len(set(nombre_db.split()) & set(maestro_planilla.split())) < min(2, len(set(nombre_db.split())))):
        return (
            "Este no es tu registro de calificaciones. "
            "El nombre del docente en la planilla no coincide con tu cuenta."
        )

    # 2. Grado
    año_planilla = _normalizar(metadatos.get('año_escolaridad', ''))
    grado_db     = _normalizar(profesor_curso.curso.grado)

    if not año_planilla:
        return "Esta plantilla no pertenece a este grado."
    if grado_db:
        if _base_ordinal(año_planilla) not in _base_ordinal(grado_db) and \
           _base_ordinal(grado_db) not in _base_ordinal(año_planilla):
            return "Esta plantilla no pertenece a este grado."

    # 3. Paralelo
    paralelo_planilla = _normalizar(metadatos.get('paralelo', ''))
    paralelo_db       = _normalizar(profesor_curso.curso.paralelo)

    if not paralelo_planilla or paralelo_planilla != paralelo_db:
        return "Esta plantilla no pertenece a este paralelo."

    # 4. Área / Materia
    area_planilla = _normalizar(metadatos.get('area', ''))
    materia_db    = _normalizar(profesor_curso.materia.nombre)

    if not area_planilla or (materia_db and not (set(area_planilla.split()) & set(materia_db.split()))):
        return "Esta no es la materia correspondiente a tu asignación."

    return None


# ── Validación de estudiantes (bidireccional) ─────────────────────────────────

def _palabras(texto):
    return set(_normalizar(texto).split())


def _coincide_nombre(palabras_a, palabras_b):
    """
    True si los conjuntos representan el mismo nombre.
    Todas las palabras del conjunto más pequeño deben estar en el mayor.
    """
    if not palabras_a or not palabras_b:
        return False
    menor = palabras_a if len(palabras_a) <= len(palabras_b) else palabras_b
    mayor = palabras_a if menor is palabras_b else palabras_b
    return menor.issubset(mayor)


def validar_estudiantes(nombres_excel, curso_id):
    """
    Validación bidireccional:
      - Cada nombre del Excel debe existir en la BD (activo o inactivo).
      - Cada estudiante activo en la BD debe estar en el Excel.
    Retorna: { es_valido, errores[], advertencias[], lista_estudiantes[], ... }
    """
    from backend.apps.students.models import Estudiante

    estudiantes_db = list(
        Estudiante.objects
        .filter(curso_id=curso_id)
        .values('nombre', 'apellido_paterno', 'apellido_materno', 'activo')
    )

    db_activos   = []
    db_inactivos = []
    for e in estudiantes_db:
        nombre_completo = f"{e['apellido_paterno']} {e['apellido_materno']} {e['nombre']}"
        entry = {
            'nombre_display': nombre_completo.strip(),
            'palabras':       _palabras(nombre_completo),
            'activo':         e['activo'],
        }
        (db_activos if e['activo'] else db_inactivos).append(entry)

    errores           = []
    advertencias      = []
    lista_estudiantes = []

    # Precomputar palabras de los nombres del Excel una sola vez
    excel_entries = [{'nombre': n, 'palabras': _palabras(n)} for n in nombres_excel]

    # Verificación 1: cada nombre del Excel debe existir en la BD
    for exc in excel_entries:
        palabras_exc   = exc['palabras']
        nombre_excel   = exc['nombre']
        match_activo   = next((e for e in db_activos   if _coincide_nombre(palabras_exc, e['palabras'])), None)
        match_inactivo = next((e for e in db_inactivos if _coincide_nombre(palabras_exc, e['palabras'])), None)

        if match_activo:
            lista_estudiantes.append({'nombre': nombre_excel, 'encontrado': True, 'activo': True})
        elif match_inactivo:
            lista_estudiantes.append({'nombre': nombre_excel, 'encontrado': True, 'activo': False})
            advertencias.append(f"El estudiante {nombre_excel} figura como inactivo en el sistema.")
        else:
            errores.append({
                'tipo':    'no_en_bd',
                'nombre':  nombre_excel,
                'mensaje': (
                    f"{nombre_excel} está en la planilla pero no en la base de datos, "
                    "pídele al director que lo agregue."
                ),
            })
            lista_estudiantes.append({'nombre': nombre_excel, 'encontrado': False, 'activo': None})

    # Verificación 2: cada activo de la BD debe estar en el Excel
    for entry in db_activos:
        en_excel = any(_coincide_nombre(exc['palabras'], entry['palabras']) for exc in excel_entries)
        if not en_excel:
            errores.append({
                'tipo':    'no_en_excel',
                'nombre':  entry['nombre_display'],
                'mensaje': (
                    f"El estudiante {entry['nombre_display']} no está en la planilla pero sí "
                    "en la base de datos, pídele al director que lo dé de baja en caso de que "
                    "no esté en la institución."
                ),
            })

    return {
        'es_valido':        len(errores) == 0,
        'errores':          errores,
        'advertencias':     advertencias,
        'lista_estudiantes': lista_estudiantes,
        'activos':          sum(1 for e in lista_estudiantes if e.get('activo') is True),
        'inactivos':        sum(1 for e in lista_estudiantes if e.get('activo') is False and e.get('encontrado') is True),
        'no_encontrados':   [e['nombre'] for e in lista_estudiantes if not e.get('encontrado')],
        'total_excel':      len(nombres_excel),
        'total_bd':         len(estudiantes_db),
    }


# ── Extracción de notas ────────────────────────────────────────────────────────

_SABER_COLS = list(range(19, 29))
_SABER_PROM = 29
_HACER_COLS = list(range(30, 40))
_HACER_PROM = 40

_FILA_TITULO_INI      = 9
_FILA_ESTUDIANTES_INI = 15
_FILA_ESTUDIANTES_FIN = 54


def _titulo_columna(ws, col_num):
    for fila in range(_FILA_TITULO_INI, 15):
        val = ws.cell(row=fila, column=col_num).value
        if val is not None and str(val).strip():
            return str(val).strip()
    return None


def _casilleros_activos(ws, columnas):
    activos = []
    for col_num in columnas:
        titulo     = _titulo_columna(ws, col_num)
        tiene_dato = False
        for fila in range(_FILA_ESTUDIANTES_INI, _FILA_ESTUDIANTES_FIN + 1):
            val = ws.cell(row=fila, column=col_num).value
            if val is not None and val != '':
                tiene_dato = True
                break
            if val == 0:
                nombre = ws.cell(row=fila, column=2).value
                if nombre and str(nombre).strip():
                    tiene_dato = True
                    break
        if titulo or tiene_dato:
            from openpyxl.utils import get_column_letter
            activos.append({
                'col':    col_num,
                'letra':  get_column_letter(col_num),
                'titulo': titulo if titulo else f"Col {get_column_letter(col_num)}",
            })
    return activos


def _estudiantes_hoja(ws, wb):
    estudiantes = []
    for fila in range(_FILA_ESTUDIANTES_INI, _FILA_ESTUDIANTES_FIN + 1):
        nombre = ws.cell(row=fila, column=2).value
        numero = ws.cell(row=fila, column=1).value
        if nombre and str(nombre).strip():
            estudiantes.append({'fila': fila, 'numero': numero, 'nombre': str(nombre).strip()})

    if not estudiantes and 'FILIACION' in wb.sheetnames:
        ws_fil = wb['FILIACION']
        offset = _FILA_ESTUDIANTES_INI - 9
        for fila in range(_FILA_ESTUDIANTES_INI, _FILA_ESTUDIANTES_FIN + 1):
            nombre = ws_fil.cell(row=fila - offset, column=2).value
            numero = ws.cell(row=fila, column=1).value
            if nombre and str(nombre).strip():
                estudiantes.append({
                    'fila':   fila,
                    'numero': numero if numero else fila - _FILA_ESTUDIANTES_INI + 1,
                    'nombre': str(nombre).strip(),
                })
    return estudiantes


def _notas_dimension(ws, estudiantes, casilleros, col_promedio):
    resultados = []
    for est in estudiantes:
        fila  = est['fila']
        notas = {}
        for cas in casilleros:
            val = ws.cell(row=fila, column=cas['col']).value
            notas[cas['titulo']] = val if val is not None else None
        promedio = ws.cell(row=fila, column=col_promedio).value
        resultados.append({
            'numero':   est['numero'],
            'nombre':   est['nombre'],
            'notas':    notas,
            'promedio': promedio,
        })
    return resultados


def extraer_notas(archivo):
    """
    Extrae las notas SABER y HACER de las hojas de trimestre.
    Acepta un Workbook ya abierto, bytes o BytesIO.
    """
    if isinstance(archivo, openpyxl.Workbook):
        wb = archivo
    else:
        if isinstance(archivo, (bytes, bytearray)):
            archivo = io.BytesIO(archivo)
        wb = openpyxl.load_workbook(archivo, data_only=True)

    _TRIM_LABELS = {
        '1TRIM': '1er Trimestre',
        '2TRIM': '2do Trimestre',
        '3TRIM': '3er Trimestre',
    }

    resultado = {'trimestres': {}}

    for sheet_name, label in _TRIM_LABELS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws            = wb[sheet_name]
        saber_activos = _casilleros_activos(ws, _SABER_COLS)
        hacer_activos = _casilleros_activos(ws, _HACER_COLS)
        estudiantes   = _estudiantes_hoja(ws, wb)

        resultado['trimestres'][sheet_name] = {
            'label': label,
            'saber': {
                'casilleros': [c['titulo'] for c in saber_activos],
                'datos':      _notas_dimension(ws, estudiantes, saber_activos, _SABER_PROM),
            },
            'hacer': {
                'casilleros': [c['titulo'] for c in hacer_activos],
                'datos':      _notas_dimension(ws, estudiantes, hacer_activos, _HACER_PROM),
            },
        }

    return resultado
